import os
import openpyxl

from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment


class ExcelModule(object):
    def __init__(self,path):
        print(self)
        self.path = path

        if(os.path.exists(self.path)):
            print(self.path,' file is exists')
        else:
            if(self.create()):
                print('create file success!')
            else:
                print('create file error!')

    def create(self):
        print(os.path.splitext(os.path.basename(self.path))[0])
        if os.path.exists(self.path):
            return False
        else:
            wb = openpyxl.Workbook()
            wb.create_sheet(index = 0, title = os.path.splitext(os.path.basename(self.path))[0])
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
            wb.save(self.path)
            return True

    def get_sheet_names(self):
        if True != os.path.exists(self.path):
            return -1

        wb = load_workbook(self.path)
        return wb.sheetnames
    
    def add_sheet_front(self, path, sheetname):
        print('add sheet front')

    def add_sheet_end(self, path, sheetname):
        print('add sheet end!')

    def read_column_data(self,sheetname,col):
        print('sheetname is [',sheetname, '] column is [',col,']')
        if True != os.path.exists(self.path):
            return -1

        wb = load_workbook(self.path)
        ws = wb[sheetname]
        
        rows = ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = ws.cell(row=i, column=col).value
            column_data.append(cell_value)
        return column_data

    def read_row_data(self,sheetname,row):
        print('sheetname is [',sheetname, '] column is [',col,']')
        if True != os.path.exists(self.path):
            return -1

        wb = load_workbook(self.path)
        ws = wb[sheetname]
        
        columns = ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data
        
    def copy_sheet(self, srcfile, tagfile ,copysheet ,sheetname):
        print('copy ',srcfile,'sheet',copysheet,'to ',tagfile)

        wb = load_workbook(srcfile,data_only=True)
        wb2 = load_workbook(tagfile)
        
        ws = wb[copysheet]
        
        ws2 = wb2.create_sheet(index = len(wb2.sheetnames), title = sheetname)
 
        #两个for循环遍历整个excel的单元格内容
        print('copying...')
        for i,row in enumerate(ws.iter_rows()):
            for j,cell in enumerate(row):
                ws2.cell(row=i+1, column=j+1, value=cell.value)
        print('copying end')
        
        wb2.save(tagfile)
        wb.close()
        wb2.close()
       
